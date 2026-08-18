"""
Reads a real, user-uploaded CDS/broker portfolio `.xlsx` export into raw
rows for `app.domain.portfolio_import_parsing` — the I/O layer that pure
module deliberately doesn't have, the same split every other real
extraction pipeline in this system draws (`app.ingestion.financial_pdf_
extractor` / `app.domain.financial_statement_parsing` is the closest
precedent).

`data_only=True` READS THE LAST-CALCULATED VALUES, NOT FORMULAS. A real
broker export's "Total Cost"/"Market Value" column are typically real
Excel formulas (`=SUM(...)`) in the underlying file, not literal numbers
— `data_only=True` reads whatever Excel itself last computed and cached
for that cell, which is what a real user opening the file in Excel
actually sees. Reading formula source text instead would hand this
module `"=SUM(J4:J12)"` rather than `75131.13`.
"""
from __future__ import annotations

import io


def read_portfolio_workbook(file_bytes: bytes) -> list[tuple]:
    """Every row of the workbook's FIRST sheet, each as a tuple of raw
    cell values in column order — exactly the shape `app.domain.
    portfolio_import_parsing.parse_portfolio_export` expects. Raises
    on anything that isn't a real, readable `.xlsx` file (not a genuine
    "no data" case — a caller should treat this as a real upload error,
    not silently return an empty portfolio)."""
    import openpyxl

    workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    return [tuple(row) for row in sheet.iter_rows(values_only=True)]
