"""Ingests the manually-downloaded CSE "Market Time & Sales History" Excel
exports in docs/price data/ into prices_daily.

Real, official, per-trade CSE data (Security, Board Id, Trade Time,
Quantity, Price, Net Change) — not a third-party aggregator. Aggregated
here into one daily OHLCV bar per (ticker, date):

    open      = price of the chronologically FIRST trade that day
    close     = price of the chronologically LAST trade that day
    high/low  = max/min trade price that day
    volume    = sum of trade quantities
    turnover  = sum of price*quantity
    vwap      = turnover / volume
    trades    = count of trade prints

REGULAR-BOARD TRADES ONLY, DELIBERATE. The `Board Id` column carries
three real values: REGULAR (continuous double-auction market), CROSSING
(privately-negotiated block deals — CSE also appends a footnote marker
directly onto the security symbol for these, e.g. "JKH.N0000C`", which
LOOKS like a distinct ticker but isn't), and AUCTION (negligible volume,
5 rows out of ~862k sampled). A single JKH crossing was seen at 2,000,000
shares in ONE print — larger than that day's entire regular-market
volume easily could be. Blending that into open/high/low/close would let
one negotiated-off-the-order-book deal set the "market" price for the
whole day, corrupting exactly the return series this data exists to
build. Filtering to REGULAR also incidentally fixes the footnote-marker
tickers: they only ever appear on non-REGULAR rows.

CONSERVATIVE MERGE POLICY, DELIBERATE: only ever INSERTS a (ticker, date)
that prices_daily does NOT already have. Never overwrites an existing
row — the existing rows already passed this system's own independent
verification (see app.ingestion.price_loader), and there is no upside
to risking a bug in THIS aggregation script silently degrading already-
trusted data. The whole point of this ingest is the OLDER depth
(pre-2025-08-19) prices_daily doesn't have at all, not a rewrite of the
recent window.

adj_factor is left at its default 1.0, matching upsert_eod_prices's own
documented convention: adjustment is rebuilt separately, for the whole
series, from corporate_actions — never written at ingest time.

source='cse.lk:market_time_sales_export' — honestly distinct from the
live API's plain 'cse.lk', since this arrived via a different, manual,
official CSE export, not the paced tradeSummary/companyChartDataByStock
endpoints.

CHECKPOINTED, DELIBERATE — a real, measured problem, not a hypothetical
one: this environment kills a long-running background process without
warning somewhere around the 20-40 minute mark (observed three times
across this session's various long-running scripts). A single "parse
all ~230 files, then one big insert at the end" run loses ALL of that
work the instant it's killed partway through. This script instead
processes files in small batches (BATCH_SIZE), commits each batch to
the DB immediately, and records every fully-committed filename to
CHECKPOINT_FILE — so a killed-and-rerun script skips straight past
whatever's already safely in the database instead of re-parsing (slow)
or re-inserting (harmless but wasteful) it.
"""
from __future__ import annotations

import datetime as dt
import glob
import sys
from decimal import Decimal

import openpyxl

sys.path.insert(0, r"C:\Users\USER\Documents\Claude Code Projects\CSE-Alpha-Project\backend")

from sqlalchemy import select

from app.db.session import SessionLocal
from app.models.prices import PriceDaily
from app.models.securities import Security

PRICE_DATA_DIR = r"C:\Users\USER\Documents\Claude Code Projects\CSE-Alpha-Project\docs\price data"
CHECKPOINT_FILE = r"C:\Users\USER\Documents\Claude Code Projects\CSE-Alpha-Project\backend\logs\manual_price_ingest_checkpoint.txt"
SOURCE_LABEL = "cse.lk:market_time_sales_export"
BATCH_SIZE = 20


class DayAgg:
    __slots__ = ("first_ts", "first_px", "last_ts", "last_px", "high", "low", "qty", "notional", "trades")

    def __init__(self) -> None:
        self.first_ts: dt.datetime | None = None
        self.first_px: float = 0.0
        self.last_ts: dt.datetime | None = None
        self.last_px: float = 0.0
        self.high: float = float("-inf")
        self.low: float = float("inf")
        self.qty: float = 0.0
        self.notional: float = 0.0
        self.trades: int = 0

    def add(self, ts: dt.datetime, price: float, qty: float) -> None:
        if self.first_ts is None or ts < self.first_ts:
            self.first_ts, self.first_px = ts, price
        if self.last_ts is None or ts >= self.last_ts:
            self.last_ts, self.last_px = ts, price
        if price > self.high:
            self.high = price
        if price < self.low:
            self.low = price
        self.qty += qty
        self.notional += price * qty
        self.trades += 1


def load_checkpoint() -> set[str]:
    try:
        with open(CHECKPOINT_FILE, encoding="utf-8") as f:
            return {line.strip() for line in f if line.strip()}
    except FileNotFoundError:
        return set()


def append_checkpoint(filenames: list[str]) -> None:
    with open(CHECKPOINT_FILE, "a", encoding="utf-8") as f:
        for name in filenames:
            f.write(name + "\n")


def parse_batch(files: list[str]) -> tuple[dict[tuple[str, dt.date], DayAgg], int, int, int]:
    agg: dict[tuple[str, dt.date], DayAgg] = {}
    total_rows = 0
    bad_rows = 0
    non_regular_rows = 0

    for f in files:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=4, values_only=True):
            if not row or row[0] is None:
                continue
            ticker, board, trade_time, qty, price = row[0], row[1], row[2], row[3], row[4]
            if board != "REGULAR":
                non_regular_rows += 1
                continue
            if ticker is None or trade_time is None or qty is None or price is None:
                bad_rows += 1
                continue
            try:
                ts = dt.datetime.strptime(trade_time, "%m/%d/%Y %H:%M:%S")
            except (ValueError, TypeError):
                bad_rows += 1
                continue
            total_rows += 1
            key = (ticker, ts.date())
            if key not in agg:
                agg[key] = DayAgg()
            agg[key].add(ts, float(price), float(qty))
        wb.close()

    return agg, total_rows, bad_rows, non_regular_rows


def ingest_batch(db, agg: dict[tuple[str, dt.date], DayAgg], known_tickers: set[str]) -> tuple[int, int, set[str]]:
    existing_dates = set(db.execute(select(PriceDaily.ticker, PriceDaily.date)).all())
    unknown_tickers: set[str] = set()
    already_present = 0
    inserted = 0
    now = dt.datetime.now(dt.timezone.utc)

    for (ticker, date), bar in agg.items():
        if ticker not in known_tickers:
            unknown_tickers.add(ticker)
            continue
        if (ticker, date) in existing_dates:
            already_present += 1
            continue
        if bar.qty <= 0:
            continue

        db.add(
            PriceDaily(
                ticker=ticker,
                date=date,
                open=Decimal(str(round(bar.first_px, 4))),
                high=Decimal(str(round(bar.high, 4))),
                low=Decimal(str(round(bar.low, 4))),
                close=Decimal(str(round(bar.last_px, 4))),
                vwap=Decimal(str(round(bar.notional / bar.qty, 4))),
                volume=int(bar.qty),
                turnover=Decimal(str(round(bar.notional, 2))),
                trades=bar.trades,
                adj_factor=Decimal("1.0"),
                fetched_at=now,
                source=SOURCE_LABEL,
            )
        )
        inserted += 1

    db.commit()
    return inserted, already_present, unknown_tickers


def main() -> None:
    all_files = sorted(glob.glob(f"{PRICE_DATA_DIR}\\*.xlsx"))
    done = load_checkpoint()
    todo = [f for f in all_files if f.split("\\")[-1] not in done]
    print(f"{len(all_files)} files total, {len(done)} already checkpointed, {len(todo)} left to process.")

    db = SessionLocal()
    known_tickers = set(db.scalars(select(Security.ticker)).all())

    total_rows = total_inserted = total_already_present = 0
    all_unknown: set[str] = set()

    for i in range(0, len(todo), BATCH_SIZE):
        batch_files = todo[i : i + BATCH_SIZE]
        agg, rows, bad, non_reg = parse_batch(batch_files)
        inserted, already_present, unknown = ingest_batch(db, agg, known_tickers)
        append_checkpoint([f.split("\\")[-1] for f in batch_files])

        total_rows += rows
        total_inserted += inserted
        total_already_present += already_present
        all_unknown |= unknown

        done_count = min(i + BATCH_SIZE, len(todo))
        print(f"  batch {done_count}/{len(todo)} files: +{rows:,} rows, +{inserted:,} inserted "
              f"(bad={bad}, non-regular={non_reg}) — committed & checkpointed")

    db.close()

    print(f"\nTOTAL this run: {total_rows:,} rows parsed, {total_inserted:,} inserted, "
          f"{total_already_present:,} already present")
    print(f"Unknown tickers this run ({len(all_unknown)}): {sorted(all_unknown)[:30]}")

    db2 = SessionLocal()
    overall = db2.execute(select(PriceDaily.date).order_by(PriceDaily.date)).scalars().first()
    overall_max = db2.execute(select(PriceDaily.date).order_by(PriceDaily.date.desc())).scalars().first()
    source_min = db2.execute(
        select(PriceDaily.date).where(PriceDaily.source == SOURCE_LABEL).order_by(PriceDaily.date)
    ).scalars().first()
    source_max = db2.execute(
        select(PriceDaily.date).where(PriceDaily.source == SOURCE_LABEL).order_by(PriceDaily.date.desc())
    ).scalars().first()
    print(f"This source's date range in prices_daily now: {source_min} -> {source_max}")
    print(f"prices_daily OVERALL date range now: {overall} -> {overall_max}")
    db2.close()


if __name__ == "__main__":
    main()
